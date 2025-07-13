import os
import openpyxl

@app.get("/merge", response_class=FileResponse)
async def handle_merge():
    """
    This new endpoint merges all .xlsx files into the template file.
    """
    template_path = os.path.join(UPLOADS_DIR, TEMPLATE_FILENAME)
    output_filename = "merged_output.xlsx"
    output_path = os.path.join(UPLOADS_DIR, output_filename)

    if not os.path.exists(template_path):
        return HTMLResponse(content=f"Template file '{TEMPLATE_FILENAME}' not found. Please upload it first.", status_code=404)

    # Load the template workbook
    merged_wb = openpyxl.load_workbook(template_path)
    merged_ws = merged_wb.active

    # Get a list of all other excel files to merge
    files_to_merge = [f for f in os.listdir(UPLOADS_DIR) if f.endswith('.xlsx') and f != TEMPLATE_FILENAME and f != output_filename]

    for filename in files_to_merge:
        filepath = os.path.join(UPLOADS_DIR, filename)
        source_wb = openpyxl.load_workbook(filepath)
        source_ws = source_wb.active

        # Copy rows from source to merged sheet, skipping the header
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            # Ensure we don't append completely empty rows
            if any(cell is not None for cell in row):
                merged_ws.append(row)

    # Save the merged workbook
    merged_wb.save(output_path)
    
    return FileResponse(path=output_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=output_filename)
