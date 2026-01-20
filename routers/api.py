import os
import shutil
import openpyxl
import json
from datetime import datetime
from fastapi import Request, UploadFile, File, APIRouter, Query, HTTPException, Depends
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from routers.authentification import verify_ip_whitelist
from routers.menu import match_agenda_user

# --- Configuration ---
UPLOADS_DIR = "uploads"
TEMPLATE_FILENAME = "template.xlsx"
MASTER_MERGE_FILENAME = "master.xlsx"
VERSIONS = ["ver1", "ver2"]
FILE_OWNERSHIP_PATH = "json/file_ownership.json"
# Keywords to match in uploaded result filenames for agenda tracking
AGENDA_KEYWORDS = ["김철수", "이영희"]

os.makedirs(UPLOADS_DIR, exist_ok=True)
for version in VERSIONS:
    os.makedirs(os.path.join(UPLOADS_DIR, version), exist_ok=True)
    os.makedirs(os.path.join(UPLOADS_DIR, version, "results"), exist_ok=True)
    os.makedirs(os.path.join(UPLOADS_DIR, version, "masterdb"), exist_ok=True)

router = APIRouter()
templates = Jinja2Templates(directory="templates")
router.mount("/static", StaticFiles(directory="static"), name="static")


# ver1 용인지 ver2 용인지 리턴
def get_version_dir(version: str):
    """Helper to get the correct versioned directory path."""
    return os.path.join(UPLOADS_DIR, version)


# --- 파일 소유권 관리 함수들 ---
def load_file_ownership():
    """파일 소유권 정보 로드"""
    if os.path.exists(FILE_OWNERSHIP_PATH):
        with open(FILE_OWNERSHIP_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_file_ownership(data):
    """파일 소유권 정보 저장"""
    with open(FILE_OWNERSHIP_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def register_file_owner(version: str, filename: str, ip: str):
    """파일 업로드 시 소유자 IP 등록 (날짜/시간 포함)"""
    ownership = load_file_ownership()
    if version not in ownership:
        ownership[version] = {}

    now = datetime.now()
    upload_info = {
        "ip": ip,
        "upload_date": now.strftime("%Y-%m-%d"),
        "upload_time": now.strftime("%H:%M:%S")
    }
    ownership[version][filename] = upload_info
    save_file_ownership(ownership)

def check_file_owner(version: str, filename: str, ip: str) -> bool:
    """현재 IP가 해당 파일의 소유자인지 확인"""
    ownership = load_file_ownership()
    file_info = ownership.get(version, {}).get(filename)
    if isinstance(file_info, dict):
        return file_info.get("ip") == ip
    # 이전 버전 호환성 (IP만 문자열로 저장된 경우)
    return file_info == ip


@router.get("/", response_class=HTMLResponse)
async def read_home(request: Request, client_ip: str = Depends(verify_ip_whitelist)):
    """
    This endpoint serves the home page.
    It now separates the template file from the data files for display.
    """
    try:
        all_files = os.listdir(UPLOADS_DIR)
        files1 = os.listdir(UPLOADS_DIR+"/ver1")
        files2 = os.listdir(UPLOADS_DIR+"/ver2")
        template_file = TEMPLATE_FILENAME if TEMPLATE_FILENAME in all_files else None
        data_files1 = [f for f in files1 if f not in [TEMPLATE_FILENAME, 'merged_output_r1.xlsx']]
        data_files2 = [f for f in files2 if f not in [TEMPLATE_FILENAME, 'merged_output_r2.xlsx']]

        # results 폴더의 파일 리스트 가져오기
        results_files1 = os.listdir(UPLOADS_DIR+"/ver1/results")
        results_files2 = os.listdir(UPLOADS_DIR+"/ver2/results")

        # masterdb 폴더의 파일 리스트 가져오기
        masterdb_files1 = os.listdir(UPLOADS_DIR+"/ver1/masterdb")
        masterdb_files2 = os.listdir(UPLOADS_DIR+"/ver2/masterdb")
    except OSError:
        template_file = None
        data_files1 = []
        data_files2 = []
        results_files1 = []
        results_files2 = []
        masterdb_files1 = []
        masterdb_files2 = []

    # 삭제 가능 여부 판단 (일반 데이터 파일만)
    deletable_files1 = {f: check_file_owner("ver1", f, client_ip) for f in data_files1}
    deletable_files2 = {f: check_file_owner("ver2", f, client_ip) for f in data_files2}

    context = {
        "request": request,
        "title": "Home Page - File Uploader",
        "message": "Upload data files below. Use the dedicated button for the template.",
        "versions": VERSIONS,
        "template_file": template_file,
        "data_files1": data_files1, # ver1에 올라간 파일 리스트
        "data_files2": data_files2, # ver2에 올라간 파일 리스트
        "results_files1": results_files1, # ver1/results에 올라간 파일 리스트
        "results_files2": results_files2, # ver2/results에 올라간 파일 리스트
        "master_merge_filename": MASTER_MERGE_FILENAME,
        "deletable_files1": deletable_files1,
        "deletable_files2": deletable_files2,
        "masterdb_files1": masterdb_files1,
        "masterdb_files2": masterdb_files2
    }
    return templates.TemplateResponse("home.html", context)


@router.post("/upload_template", response_class=RedirectResponse)
async def handle_upload_template(
    file: UploadFile = File(...),
    client_ip: str = Depends(verify_ip_whitelist)
):
    """
    NEW: Dedicated endpoint for uploading the template.xlsx file.
    It will always be saved as 'template.xlsx'.
    """
    file_path = os.path.join(UPLOADS_DIR, TEMPLATE_FILENAME)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return RedirectResponse(url="/", status_code=303)


@router.post("/upload_result/{version}", response_class=RedirectResponse)
async def upload_result(
    version: str,
    file: UploadFile = File(...),
    client_ip: str = Depends(verify_ip_whitelist)
):
    """결과 파일 올리기"""
    now = datetime.now()
    timestamp = now.strftime("%y%m%d_%H시%M분")
    filename = f"result_{version}_{timestamp}.xlsx"
    results_dir = os.path.join(get_version_dir(version), "results")
    file_path = os.path.join(results_dir, filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return RedirectResponse(url="/", status_code=303)


@router.post("/upload/{version}", response_class=RedirectResponse)
async def handle_upload(
    version: str,
    file: UploadFile = File(...),
    client_ip: str = Depends(verify_ip_whitelist)
):
    file_path = os.path.join(get_version_dir(version), file.filename)
    # Prevent overwriting the template with a data file of the same name
    if file.filename == TEMPLATE_FILENAME:
        return HTMLResponse(content="Cannot upload a data file with the name 'template.xlsx'. Please use the dedicated template upload button.", status_code=400)

    # 임시로 파일 저장
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # A5 셀 검증
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        a5_value = ws['A5'].value
        wb.close()

        # A5 셀이 비어있거나 None이면 업로드 거부
        if a5_value is None or str(a5_value).strip() == "":
            os.remove(file_path)  # 업로드된 파일 삭제
            return HTMLResponse(content="DRM을 해제해 평문으로 올려주세요", status_code=400)
    except Exception as e:
        # 파일 읽기 오류 시 파일 삭제
        if os.path.exists(file_path):
            os.remove(file_path)
        return HTMLResponse(content=f"파일을 읽는 중 오류가 발생했습니다: {str(e)}", status_code=400)

    # A5 셀에 데이터가 있으면 IP와 업로드 정보 등록
    register_file_owner(version, file.filename, client_ip)

    # Check if filename contains any agenda keywords and update agenda_no.json
    match_agenda_user(file.filename, version, file_path, AGENDA_KEYWORDS)

    return RedirectResponse(url="/", status_code=303)


# 마스터 DB 업로드
@router.post("/upload/{version}/masterdb", response_class=RedirectResponse)
async def upload_masterdb(
    version: str,
    file: UploadFile = File(...),
):
    masterdb_dir = os.path.join(get_version_dir(version), "masterdb")
    file_path = os.path.join(masterdb_dir, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return RedirectResponse(url="/", status_code=303)


@router.get("/download/{version}/{filename:path}", response_class=FileResponse)
async def handle_download(
    version: str,
    filename: str,
    client_ip: str = Depends(verify_ip_whitelist)
):
    file_path = os.path.join(get_version_dir(version), filename)
    if os.path.exists(file_path):
        # 실제 파일명만 추출 (경로 제외)
        actual_filename = os.path.basename(filename)
        return FileResponse(path=file_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=actual_filename)
    return HTMLResponse(content="File not found.", status_code=404)



@router.get("/delete/{version}/{filename:path}", response_class=RedirectResponse)
async def handle_delete(
    version: str,
    filename: str,
    client_ip: str = Depends(verify_ip_whitelist)
):
    if ".." in filename or filename.startswith("/"):
        return HTMLResponse(content="Invalid filename.", status_code=400)

    # results 폴더가 아닌 일반 파일의 경우 권한 검증
    if not filename.startswith("results/"):
        if not check_file_owner(version, filename, client_ip):
            return HTMLResponse(content="삭제 권한이 없습니다. 본인이 업로드한 파일만 삭제할 수 있습니다.", status_code=403)

    file_path = os.path.join(get_version_dir(version), filename)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except OSError as e:
            print(f"Error deleting file {filename}: {e}")
    return RedirectResponse(url="/", status_code=303)


@router.get("/merge/{version}", response_class=FileResponse)
async def handle_merge(
    version: str,
    client_ip: str = Depends(verify_ip_whitelist)
):
    """
    uploads/template.xlsx를 베이스로 사용하여
    uploads/{version} 내의 모든 엑셀 파일을 하나로 합치기
    - 각 파일의 5번째 행부터 데이터 가져오기 (1-4행 스킵)
    - template.xlsx의 5번째 행부터 데이터 붙여넣기
    - 각 행의 마지막 데이터가 있는 열까지만 복사
    """
    # template.xlsx 확인
    template_path = os.path.join(UPLOADS_DIR, TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="template.xlsx 파일이 없습니다.")

    # 현재 시간으로 파일명 생성
    now = datetime.now()
    timestamp = now.strftime("%y%m%d_%H_%M")
    output_filename = f"merged_output_{version}_{timestamp}.xlsx"
    output_path = os.path.join(get_version_dir(version)+"/mergedoutput", output_filename)

    # template.xlsx를 베이스로 워크북 로드
    merged_wb = openpyxl.load_workbook(template_path)
    merged_ws = merged_wb.active

    # uploads/{version} 폴더의 모든 xlsx 파일 가져오기
    files_to_merge = [f for f in os.listdir(get_version_dir(version))
                      if f.endswith('.xlsx') and f != output_filename
                      and not f.startswith('merged_output_')]

    # 현재 붙여넣기를 시작할 행 번호 (5번째 행부터 시작)
    current_row = 5

    # 각 파일을 순회하며 데이터 복사
    for filename in files_to_merge:
        filepath = os.path.join(get_version_dir(version), filename)
        source_wb = openpyxl.load_workbook(filepath)
        source_ws = source_wb.active

        # 5번째 행부터 데이터 읽기
        for row in source_ws.iter_rows(min_row=5, max_col=11, values_only=True):
            # 행에 데이터가 있는지 확인
            if any(cell is not None for cell in row):
                # 뒤에서부터 확인해서 마지막 데이터가 있는 열 찾기
                last_data_idx = None
                for i in range(len(row) - 1, -1, -1):
                    if row[i] is not None:
                        last_data_idx = i
                        break

                # 마지막 데이터가 있는 열까지만 복사
                if last_data_idx is not None:
                    for col_idx, cell_value in enumerate(row[:last_data_idx + 1], start=1):
                        merged_ws.cell(row=current_row, column=col_idx, value=cell_value)
                    current_row += 1

        source_wb.close()

    # 병합된 파일 저장
    merged_wb.save(output_path)
    merged_wb.close()

    return FileResponse(path=output_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=output_filename)
    

@router.get("/detail", response_class=HTMLResponse)
async def read_about(request: Request):
    return OSError